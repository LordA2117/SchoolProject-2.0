from openpyxl import load_workbook
from os import listdir
from pprint import pprint


def readWorkbook(path) -> dict:
    wb = load_workbook(path)
    ws = wb.active
    num = ws.max_row
    compiled = {}
    for i in range(2, num + 1):
        lis = [i.value for i in ws[i]]
        compiled[lis[0]] = lis[1:]
    return compiled


# pprint(readWorkbook(
#     r"C:\Users\abhin\OneDrive\Desktop\SchoolData\Bhaswati Chattopadhyay.xlsx")
# )

# 2 kinds of clashes:
# 1. Both teachers assigned to same period
# 2. One teacher assigned to 2 places at once


def checkClashes(path1, path2, day) -> list:
    # Checks clashes if both teachers are assigned to one period, accounting for the fact that they teach different streams
    t1 = readWorkbook(path1)
    t2 = readWorkbook(path2)
    day1 = t1[day]
    day2 = t2[day]
    clashes = []
    for i in range(len(day1)):
        if day1[i] and day2[i]:
            p1 = day1[i].split('\n')
            p2 = day2[i].split('\n')
            for j in p1:
                if j in p2:
                    clashes.append(
                        f'Period {i + 1}: Clash detected. 2 teachers assigned in the same period for 1 or more groups.')
                    break

    return clashes


# res = checkClashes(r"C:\Users\abhin\OneDrive\Desktop\SchoolData\Bini P Kuriakose.xlsx",
#                    r"C:\Users\abhin\OneDrive\Desktop\SchoolData\Bini P Kuriakose.xlsx", 'Tuesday')

# pprint(res)


def viewFreeAndBusy(folder, day, period_number, view_busy=False) -> list:
    files = listdir(folder)
    freeTeachers = []
    busyTeachers = []
    for i in files:
        period = readWorkbook(f'{folder}\{i}')[
            day][period_number - 1]
        if period == None:
            freeTeachers.append(i)
        else:
            busyTeachers.append(i)

    if view_busy == False:
        return freeTeachers
    return busyTeachers
